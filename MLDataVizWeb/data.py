import numpy as np
import csv
from openpyxl import Workbook

# Dosyayı oku ve verileri al
with open("data/exercises4allparticipantsCOP.txt", "r") as file:
    lines = file.readlines()

# Verileri saklamak için bir liste oluştur
column_6_data = []
column_7_data = []

# Her satırı ayır ve istenen sütunlardaki verileri al
for line in lines:
    parts = line.split()
    if len(parts) < 7:
        continue
    # Sütun verilerini al ve listeye ekle
    column_6_data.append(float(parts[5]))  # 0'dan başladığı için 6. sütun 5. sütun
    column_7_data.append(float(parts[6]))  # 0'dan başladığı için 7. sütun 6. sütun

# 6. ve 7. sütunların ortalamalarını hesapla
column_6_mean = np.mean(column_6_data)
column_7_mean = np.mean(column_7_data)

# Her bir değerden ortalamayı çıkar ve yeni veri setini oluştur
new_column_6_values = [value - column_6_mean for value in column_6_data]
new_column_7_values = [value - column_7_mean for value in column_7_data]

# Excel dosyası oluştur
wb = Workbook()
ws = wb.active

# Verileri Excel dosyasına yaz
for i in range(len(new_column_6_values)):
    ws.cell(row=i+1, column=1, value=new_column_6_values[i])  # 6. sütun
    ws.cell(row=i+1, column=2, value=new_column_7_values[i])  # 7. sütun

# Excel dosyasını kaydet
wb.save("Exercise4Complete.xlsx")
